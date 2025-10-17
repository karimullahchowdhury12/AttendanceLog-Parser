import os
import json
import csv
from datetime import datetime, time
from collections import defaultdict
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment

SHIFT_START_TIME = time(9,0)
SHIFT_END_TIME = time(18,0)
LATE_THRESHOLD = time(9,30)
EARLY_THRESHOLD = time(17,0)
LOG_FOLDER = 'attendance_logs'
OUTPUT_FOLDER = 'attendance_reports'


def read_log_files(log_folder: str) -> Tuple[defaultdict, List[str], set]:
    attendance_date = defaultdict(lambda: defaultdict(list))
    error_log = []
    processed_records = set()

    if not os.path.exists(log_folder):
        error_log.append(f"Error folder '{log_folder}' does not exist.")
        return attendance_date, error_log, processed_records

    files = [f for f in os.listdir(log_folder) if f.endswith(('.log', '.csv'))]

    if not files:
        error_log.append(f"No log files found in '{log_folder}'.")
        return attendance_date, error_log, processed_records

    for file in files:
        filepath = os.path.join(log_folder, file)
        process_file(filepath, attendance_date, error_log, processed_records)

    return attendance_date, error_log, processed_records


def process_file(filepath:str, attendance_data: defaultdict, error_log: List[str], processed_records: set) -> None:
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            first_line = f.readline().strip()
            f.seek(0)

            if ',' in first_line:
                render = csv.DictReader(f)
                for row_num, row in enumerate(render, start=2):
                    process_row(row, filepath, row_num, attendance_data, error_log, processed_records)
            else:
                first_line_values = first_line.split()
                is_header = first_line_values[0] in ['emp_code', 'employee_code', 'code']

                row_num = 1
                for line in f:
                    row_num += 1
                    line = line.strip()

                    if not line or line.startswith('#'):
                        continue

                    if row_num == 2 and is_header:
                        continue

                    values = line.split()
                    if len(values) < 6:
                        error_log.append(f"Row {row_num-1} in file '{filepath}' has insufficient columns.")
                        continue

                    row = {
                        'emp_code': values[0],
                        'first_name': values[1],
                        'last_name': values[2],
                        'timestamp': values[3],
                        'device': ' '.join(values[4:])
                    }
                    process_row(row, filepath, row_num, attendance_data, error_log, processed_records)

    except Exception as e:
        error_log.append(f"Error processing file '{filepath}': {str(e)}")


def process_row(row: Dict, filepath: str, row_num: int, attendance_data: defaultdict, error_log: List[str], processed_records: set) -> None:
    try:
        emp_code = row.get('emp_code', '').strip()
        first_name = row.get('first_name', '').strip()
        last_name = row.get('last_name', '').strip()
        timestamp = row.get('timestamp', '').strip()
        device = row.get('device', '').strip()

        if not emp_code or not first_name or not last_name or not timestamp or not device:
            error_log.append(f"Row {row_num} in file '{filepath}' is missing required fields.")
            return

        if not emp_code.isalnum():
            error_log.append(f"Row {row_num} in file '{filepath}'emp_code must be alphanumeric.")
            return
        if not first_name.isalpha():
            error_log.append(f"Row {row_num} in file '{filepath}' first_name must contain character.")
            return
        if not last_name.isalpha():
            error_log.append(f"Row {row_num} in file '{filepath}' last_name must contain character.")
            return
        if not timestamp.isdigit() and not any(c in timestamp for c in ['-', '/', ':']):
            error_log.append(f"Row {row_num} in file '{filepath}' timestamp format is invalid.")
            return
        if not device:
            error_log.append(f"Row {row_num} in file '{filepath}' device format is invalid or missing.")


        dt = parse_timestamp(timestamp)
        if dt is None:
            error_log.append(f"Row {row_num} in file '{filepath}' has an invalid timestamp.")
            return

        record_id = (emp_code, dt.timestamp(), device)
        if record_id in processed_records:
            return

        processed_records.add(record_id)

        date_key = dt.date().isoformat()
        attendance_data[emp_code][date_key].append(dt)

    except Exception as e:
        error_log.append(f"Error processing row {row_num} in file '{filepath}': {str(e)}")

def parse_timestamp(timestamp:str) -> datetime:
    try:
        if timestamp.isdigit():
            return datetime.fromtimestamp(int(timestamp))

        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y %H:%M',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%Y/%m/%d %H:%M:%S',
            '%Y/%m/%d %H:%M'
        ]

        for fmt in formats:
            try:
                return datetime.strptime(timestamp, fmt)
            except ValueError:
                continue
        return None
    except:
        return None


def calculate_summary(attendance_date: defaultdict) -> Dict:
    summary = defaultdict(list)

    for emp_code in sorted(attendance_date.keys()):
        emp_dates = attendance_date[emp_code]

        for date in sorted(emp_dates.keys()):
            punches = sorted(emp_dates[date])

            if not punches:
                continue

            first_punch = punches[0]
            last_punch = punches[-1]
            total_punches = len(punches)

            duration = last_punch - first_punch
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            work_hours = f"{hours:02d}:{minutes:02d}"

            late_entry = 1 if first_punch.time() > LATE_THRESHOLD else 0
            early_exit = 1 if last_punch.time() < EARLY_THRESHOLD else 0
            single_punch = 1 if total_punches == 1 else 0

            summary[date].append({
                'emp_code': emp_code,
                'first_punch': first_punch.strftime('%H:%M'),
                'last_punch': last_punch.strftime('%H:%M'),
                'total_punches': total_punches,
                'working_hours': work_hours,
                'late_entry': late_entry,
                'early_exit': early_exit,
                'single_punch': single_punch
            })

    for date in summary:
        summary[date].sort(key=lambda x: x['emp_code'])

    return dict(summary)

def save_json_summary(summary: Dict, output_folder: str = OUTPUT_FOLDER, output_file: str='attendance_summery.json') -> None:
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, output_file)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=4)
        print(f"Summary saved to '{filepath}'.")

def save_excel_summary(summary: Dict, output_folder: str = OUTPUT_FOLDER, output_file: str='attendance_summery.xlsx') -> None:
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, output_file)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Summary'

    header = ['Date', 'Emp Code', 'First Punch', 'Last Punch', 'Total Punches', 'Working Hours', 'Late Entry', 'Early Exit', 'Single Punch']

    header_font = Font(bold=True, color='000000', size=12)

    for col, header in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for date in sorted(summary.keys()):
        for record in summary[date]:
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=record['emp_code'])
            ws.cell(row=row, column=3, value=record['first_punch'])
            ws.cell(row=row, column=4, value=record['last_punch'])
            ws.cell(row=row, column=5, value=record['total_punches'])
            ws.cell(row=row, column=6, value=record['working_hours'])
            ws.cell(row=row, column=7, value='YES' if record['late_entry'] else 'No')
            ws.cell(row=row, column=8, value='YES' if record['early_exit'] else 'No')
            ws.cell(row=row, column=9, value='YES' if record['single_punch'] else 'No')
            row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except:
                pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filepath)

def save_error_log(error_log: List[str], output_folder: str = OUTPUT_FOLDER, output_file: str='error_log.txt') -> None:
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, output_file)
    with open(filepath, 'w', encoding='utf-8') as f:
        if error_log:
            f.write(f"Total Errors: {len(error_log)}\n")
            f.write(("=" * 80 + "\n\n"))
            for error in error_log:
                f.write(error + "\n")
        else:
            f.write("No errors found.")
    print(f"Error log saved to '{filepath}'.")


def search_attendance(attendance_data: defaultdict, emp_code:str, date:str = None) -> List[Dict]:
    if emp_code not in attendance_data:
        return []

    records = []
    emp_dates = attendance_data[emp_code]

    dates_to_search = [date] if date else sorted(emp_dates.keys())

    for search_date in dates_to_search:
        if search_date in emp_dates:
            punches = sorted(emp_dates[search_date])
            if punches:
                first_punch = punches[0]
                last_punch = punches[-1]

                duration = last_punch - first_punch
                hours = duration.seconds / 3600
                minutes = (duration.seconds % 3600) / 60
                work_hours = round(hours, 2)

                records.append({
                    'emp_code': emp_code,
                    'first_punch': first_punch.strftime('%H:%M'),
                    'last_punch': last_punch.strftime('%H:%M'),
                    'total_punches': len(punches),
                    'working_hours': work_hours,
                    'late_entry': 1 if first_punch.time() > LATE_THRESHOLD else 0,
                    'early_exit': 1 if last_punch.time() < EARLY_THRESHOLD else 0
                })

    return records

def search_summary_by_employee(summary: Dict, emp_code: str) -> List[Dict]:
    results = []
    for date in sorted(summary.keys()):
        for record in summary[date]:
            if record['emp_code'] == emp_code:
                results.append({
                    'date': date,
                    **record
                })
    return results

def search_summary_by_date(summary: Dict, date: str) -> List[Dict]:
    if date not in summary:
        return []
    return sorted(summary[date], key=lambda x: x['emp_code'])

def search_summary_by_employee_and_date(summary: Dict, emp_code: str, date: str) -> Dict:
    if date not in summary:
        return {}
    for record in summary[date]:
        if record['emp_code'] == emp_code:
            return {
                'date': date,
                **record
            }
    return {}

def search_summary_by_date_range(summary: Dict, emp_code: str, start_date: str, end_date: str) -> List[Dict]:
    results = []
    for date in sorted(summary.keys()):
        if start_date <= date <= end_date:
            for record in summary[date]:
                if record['emp_code'] == emp_code:
                    results.append({
                        'date': date,
                        **record
                    })
    return results

def export_search_results_json(results: List[Dict], output_folder: str = OUTPUT_FOLDER, output_file: str='search_results.json') -> None:
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, output_file)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4)

def export_search_results_excel(results: List[Dict], output_folder: str = OUTPUT_FOLDER, output_file: str='search_results.xlsx') -> None:
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, output_file)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Search Results'

    if not results:
        ws.cell(row=1, column=1, value='No results found.')
        wb.save(filepath)
        print(f"Search results saved to '{filepath}'.")
        return

    header = ['Date', 'Emp Code', 'First Punch', 'Last Punch', 'Total Punches', 'Working Hours', 'Late Entry', 'Early Exit', 'Single Punch']
    header_font = Font(bold=True, color='000000', size=12)

    for col, header_text in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 2
    for record in results:
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('emp_code', ''))
        ws.cell(row=row, column=3, value=record.get('first_punch', ''))
        ws.cell(row=row, column=4, value=record.get('last_punch', ''))
        ws.cell(row=row, column=5, value=record.get('total_punches', ''))
        ws.cell(row=row, column=6, value=record.get('working_hours', ''))
        ws.cell(row=row, column=7, value='YES' if record.get('late_entry', 0) else 'No')
        ws.cell(row=row, column=8, value='YES' if record.get('early_exit', 0) else 'No')
        ws.cell(row=row, column=9, value='YES' if record.get('single_punch', 0) else 'No')
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filepath)

def search_by_employee_code(summary: Dict, emp_code: str) -> None:
    results = search_summary_by_employee(summary, emp_code)
    if results:
        export_search_results_json(results, output_file=f'search_emp_{emp_code}.json')
        export_search_results_excel(results, output_file=f'search_emp_{emp_code}.xlsx')
    else:
        print(f"No results found for employee code '{emp_code}'.")

def search_by_date(summary: Dict, date: str) -> None:
    results = search_summary_by_date(summary, date)
    if results:
        export_search_results_json(results, output_file=f'search_date_{date}.json')
        export_search_results_excel(results, output_file=f'search_date_{date}.xlsx')
    else:
        print(f"No result found for date '{date}'.")

def search_by_employee_and_date(summary: Dict, emp_code: str, date: str) -> None:
    result = search_summary_by_employee_and_date(summary, emp_code, date)
    if result:
        export_search_results_json([result], output_file=f'search_emp_{emp_code}_date_{date}.json')
        export_search_results_excel([result], output_file=f'search_emp_{emp_code}_date_{date}.xlsx')
    else:
        print(f"No results found for employee code '{emp_code}' and date '{date}'.")

def search_by_date_range(summary: Dict, emp_code: str, start_date: str, end_date: str) -> None:
    results = search_summary_by_date_range(summary, emp_code, start_date, end_date)
    if results:
        export_search_results_json(results, output_file=f'search_emp_{emp_code}_{start_date}_to_{end_date}.json')
        export_search_results_excel(results, output_file=f'search_emp_{emp_code}_{start_date}_to_{end_date}.xlsx')
    else:
        print(f"No results found for employee {emp_code} from {start_date} to {end_date}")

def unix_to_date(unix_timestamp: str) -> str:
    try:
        timestamp_int = int(unix_timestamp)
        if 946684800 <= timestamp_int <= 4102444800:
            dt = datetime.fromtimestamp(timestamp_int)
            return dt.date().isoformat()
    except:
        pass
    return unix_timestamp

def parse_arguments():
    import sys
    args = sys.argv[1:]

    if not args:
        return None

    if args[0] == '--search':
        if len(args) < 2:
            print("Error: --search requires a search type.")
            return None

        search_type = args[1]

        if search_type == 'employee':
            if len(args) < 3:
                print("Error: --search employee requires an employee code.")
                return None
            emp_code = args[2]
            return ('employee', emp_code)

        elif search_type == 'date':
            if len(args) < 3:
                print("Error: --search date requires a date (YYYY-MM-DD or Unix timestamp).")
                return None
            date_input = args[2]
            date = unix_to_date(date_input)
            return ('date', date)

        elif search_type == 'employee_and_date':
            if len(args) < 4:
                print("Error: --search employee_and_date requires an employee code and a date (YYYY-MM-DD or Unix timestamp).")
                return None
            emp_code = args[2]
            date_input = args[3]
            date = unix_to_date(date_input)
            return ('employee_and_date', emp_code, date)

        elif search_type == 'date_range':
            if len(args) < 5:
                print("Error: --search date_range requires an employee code, start date, and end date (YYY-MM-DD).")
                return None
            emp_code = args[2]
            start_date_input = args[3]
            end_date_input = args[4]
            start_date = unix_to_date(start_date_input)
            end_date = unix_to_date(end_date_input)
            return ('date_range', emp_code, start_date, end_date)
        else:
            print(f"Error: Invalid search type '{search_type}'.")
            print("Valid search types: employee, date, employee_and_date, date_range")
            return None
    else:
        print(f"Error: Invalid search option '{args[0]}'.")
        print("Usage: ")
        print("python process_attendance.py --search employee <emp_code>")
        print("python process_attendance.py --search date <date>")
        print("python process_attendance.py --search employee_and_date <emp_code> <date>")
        print("python process_attendance.py --search date_range <emp_code> <start_date> <end_date>")
        return None

def process_attendance():
    #Read Data
    attendance_data, error_log, processed_records = read_log_files(LOG_FOLDER)
    if not attendance_data:
        save_error_log(error_log)
        return

    #Calculate Summary
    summary = calculate_summary(attendance_data)

    #Save all results
    save_json_summary(summary)
    save_excel_summary(summary)
    save_error_log(error_log)

    #Search Function
    search_params = parse_arguments()

    if search_params is None:
        print("\nDefault report generated. No specific search performed.")
    elif search_params[0] == 'employee':
        search_by_employee_code(summary, search_params[1])
    elif search_params[0] == 'date':
        search_by_date(summary, search_params[1])
    elif search_params[0] == 'employee_and_date':
        search_by_employee_and_date(summary, search_params[1], search_params[2])
    elif search_params[0] == 'date_range':
        search_by_date_range(summary, search_params[1], search_params[2], search_params[3])
    else:
        print("Invalid search option.")


if __name__ == "__main__":
    process_attendance()
